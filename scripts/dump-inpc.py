# Extrae la tabla de INPC y la fórmula exacta de renovación.
import openpyxl, os, sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
DL = r"C:\Users\harro\Downloads"
path = os.path.join(DL, [f for f in os.listdir(DL) if f.startswith("RENOVACIONES CALCULA")][0])

# Valores (INPC table)
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
ws = wb["Rena-INC (2)"]
print("=== TABLA INPC (año | mes | valor) — sheet 'Rena-INC (2)' ===")
n = 0
for row in ws.iter_rows(values_only=True):
    yr, mes, val = row[3], row[4], row[5]
    if mes and val is not None:
        print(f"  {str(yr) if yr else '   '} | {str(mes)[:10]:10} | {val}")
        n += 1
    if n > 130: break
wb.close()

# Fórmulas (cómo calcula la renovación)
print("\n=== FÓRMULAS sheet 'CALCULADORA' (col B y D-I, filas 5-22) ===")
wbf = openpyxl.load_workbook(path, read_only=False, data_only=False)
wf = wbf["CALCULADORA"]
for r in range(3, 23):
    for c in range(2, 10):
        cell = wf.cell(row=r, column=c)
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            print(f"  {cell.coordinate}: {v}")
wbf.close()
