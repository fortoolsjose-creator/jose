# Vuelca los Excels nuevos de la junta con Magaly.
import openpyxl, os, sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
DL = r"C:\Users\harro\Downloads"

targets = ["RENOVACIONES CALCULADORA", "SEGUIMIENTO RENTAS", "FONDO DE MANTENIMIENTO 2026 (pendiente ajuste-) (1)"]

def cell(v):
    if v is None: return ""
    if isinstance(v, float): return f"{v:,.2f}".rstrip("0").rstrip(".")
    return str(v).replace("\n", " ").strip()[:30]

for t in targets:
    matches = [f for f in os.listdir(DL) if f.startswith(t[:20]) and f.endswith(".xlsx")]
    if not matches:
        print(f"\n### NO ENCONTRADO: {t}")
        continue
    path = os.path.join(DL, sorted(matches)[-1])
    print("\n" + "="*95)
    print("ARCHIVO:", os.path.basename(path))
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("HOJAS:", wb.sheetnames)
    for ws in wb.worksheets:
        print(f"\n--- HOJA '{ws.title}' ({ws.max_row}x{ws.max_column}) ---")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = [cell(c) for c in row[:12]]
            if any(vals):
                print(f"{i:>3}|" + " | ".join(vals))
            if i > 35:
                print("   ...(más filas)")
                break
    wb.close()
