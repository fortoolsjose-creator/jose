# Vuelca la hoja CONCENTRADO del archivo de fondo de mantenimiento.
import openpyxl

PATH = r"C:\Users\harro\Downloads\FONDO DE MANTENIMIENTO 2026 (pendiente ajuste-).xlsx"
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print("HOJAS:", wb.sheetnames, "\n")

def c(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:,.0f}"
    return str(v).replace("\n", " ").strip()[:24]

for name in wb.sheetnames:
    if "CONCENT" not in name.upper():
        continue
    ws = wb[name]
    print(f"=== {name} ({ws.max_row}x{ws.max_column}) ===")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [c(x) for x in row[:10]]
        if any(vals):
            print(f"{i:>3} | " + " | ".join(vals))
        if i > 60:
            break
