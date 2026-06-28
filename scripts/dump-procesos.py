# Vuelca la estructura del archivo de procesos de operación de la asistente.
import openpyxl, os, sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
DL = r"C:\Users\harro\Downloads"
_m = [f for f in os.listdir(DL) if f.startswith("Procesos") and f.endswith(".xlsx")]
PATH = os.path.join(DL, _m[0])

def cell(v):
    if v is None:
        return ""
    s = str(v).replace("\n", " ").strip()
    return s[:40]

print("ARCHIVO:", os.path.basename(PATH))
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print("HOJAS:", wb.sheetnames, "\n")
for ws in wb.worksheets:
    print(f"\n{'='*90}\nHOJA: '{ws.title}'  ({ws.max_row} filas x {ws.max_column} cols)")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [cell(c) for c in row[:10]]
        if any(vals):
            print(f"{i:>3} | " + " | ".join(vals))
        if i > 60:
            print("   ... (truncado)")
            break
wb.close()
