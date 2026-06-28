import openpyxl, os, sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
DL = r"C:\Users\harro\Downloads"

def find(prefix):
    ms = [f for f in os.listdir(DL) if f.startswith(prefix) and f.endswith(".xlsx")]
    return os.path.join(DL, sorted(ms)[-1]) if ms else None

def cell(v):
    if v is None: return ""
    if isinstance(v, float): return f"{v:,.2f}".rstrip("0").rstrip(".")
    return str(v).replace("\n", " ").strip()[:26]

def dump(path, sheets=None, rows=16):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("\n" + "="*95 + f"\n{os.path.basename(path)}\nHOJAS: {wb.sheetnames}")
    for ws in wb.worksheets:
        if sheets and ws.title not in sheets: continue
        print(f"\n--- '{ws.title}' ({ws.max_row}x{ws.max_column}) ---")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = [cell(c) for c in row[:11]]
            if any(vals): print(f"{i:>3}|" + " | ".join(vals))
            if i > rows: break
    wb.close()

f = find("FONDO DE MANTENIMIENTO 2026 (pendiente ajuste-) (1)") or find("FONDO DE MANTENIMIENTO 2026 (pendiente")
if f: dump(f)

s = find("SEGUIMIENTO RENTAS")
if s:
    dump(s, sheets=["EGRESOS","INGRESOS","CALENDARIO DE RENOVACIONES","RENOVACIONES LISTADO","RESUMEN GENERAL","MANTENIMIENTOS","CONCILIACIÓN"], rows=20)
