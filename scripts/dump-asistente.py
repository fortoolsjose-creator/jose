# Vuelca la estructura (hojas + primeras filas) de los Excels de la asistente.
import openpyxl, os

FILES = [
    r"C:\Users\harro\Downloads\_docs_asistente\Administracion\Base de datos.xlsx",
    r"C:\Users\harro\Downloads\_docs_asistente\Administracion\METROS REDONDOS ADMÓN BASE.xlsx",
]

def cell(v):
    if v is None:
        return ""
    s = str(v).replace("\n", " ").strip()
    return s[:22]

for path in FILES:
    print("\n" + "=" * 90)
    print("ARCHIVO:", os.path.basename(path))
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print("  ERROR:", e)
        continue
    for ws in wb.worksheets:
        print(f"\n  HOJA: '{ws.title}'  ({ws.max_row} filas x {ws.max_column} cols)")
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 8:
                break
            vals = [cell(c) for c in row[:16]]
            # quita filas totalmente vacías
            if any(vals):
                rows.append(vals)
        for r in rows:
            print("    | " + " | ".join(r))
    wb.close()
